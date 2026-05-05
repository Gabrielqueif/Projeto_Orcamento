"""
Parser de planilhas Excel da SINAPI.

Encapsula a lógica de detecção de cabeçalhos, classificação de abas
e extração de registros de composições e preços por estado.
"""

import logging
import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Dict, List, Optional, Tuple, Any

import pandas as pd
from openpyxl import load_workbook
from .base_excel_parser import BaseExcelParser

from app.services.sinapi_text_utils import (
    limpar_valor_moeda,
    normalizar_nome_aba,
    remover_acentos,
)

logger = logging.getLogger(__name__)

# Siglas dos 27 estados brasileiros (incluindo DF)
COLUNAS_ESTADOS = [
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO",
    "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI",
    "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO",
]

# Abas que devem ser ignoradas no processamento
_ABAS_IGNORADAS = {"MENU", "BUSCA", "OBS", "INSTRUCOES"}

# Palavras-chave para identificar abas de preços de composições
_TERMOS_PRECOS = {"csd", "ccd", "cse"}

# Palavras-chave para identificar abas de preços de insumos
_TERMOS_INSUMOS = {"isd", "icd", "ise"}

# Palavras-chave para identificar a aba analítica
_TERMOS_ANALITICO = {"analitico", "analítico"}

# Mapeamento de palavras-chave no nome da aba → tipo de composição
_MAPA_TIPOS = {
    "CCD": "Com Desoneração",
    "COM DESONERACAO": "Com Desoneração",
    "CSE": "Empreitada",
    "EMPREITADA": "Empreitada",
    "CSD": "Sem Desoneração",
    "SEM DESONERACAO": "Sem Desoneração",
}


@dataclass
class HeaderInfo:
    """Informações da linha de cabeçalho detectada em uma aba."""

    linha: int
    """Índice da linha do cabeçalho no DataFrame."""

    col_codigo: int = -1
    """Índice da coluna CODIGO."""

    col_descricao: int = -1
    """Índice da coluna DESCRICAO."""

    col_unidade: int = -1
    """Índice da coluna UNIDADE."""

    col_grupo: int = -1
    """Índice da coluna GRUPO."""

    mapa_estados: Dict[int, str] = field(default_factory=dict)
    """Mapeia índice da coluna → sigla do estado (lowercase)."""


class SinapiExcelParser(BaseExcelParser):
    """Parser para arquivos Excel no formato SINAPI.

    Uso típico:
        >>> parser = SinapiExcelParser(file_content)
        >>> abas = parser.identificar_abas_dados()
        >>> for aba in abas:
        ...     composicoes, precos = parser.extrair_registros_aba(aba, "09/2025")
    """

    def __init__(self, file_content: bytes):
        super().__init__(file_content)

    @property
    def sheet_names(self) -> List[str]:
        """Nomes de todas as abas do arquivo."""
        return self._xl.sheet_names

    def ler_aba(self, sheet_name: str, **kwargs) -> pd.DataFrame:
        """Lê uma aba do arquivo Excel como DataFrame."""
        return pd.read_excel(self._xl, sheet_name=sheet_name, **kwargs)

    # ------------------------------------------------------------------
    # Identificação de abas
    # ------------------------------------------------------------------

    def identificar_abas_dados(self) -> List[str]:
        """Identifica quais abas contêm dados de preços de composições."""
        return self.identificar_abas_precos()

    def identificar_aba_analitico(self) -> Optional[str]:
        """Identifica a aba Analítico (hierárquia pai→filho).
        Prioriza abas sem 'custo' no nome (o Análitico puro, não o 'com custo').
        """
        candidatos = [
            a for a in self._xl.sheet_names
            if any(t in normalizar_nome_aba(a) for t in _TERMOS_ANALITICO)
        ]
        # Prefere o Analítico puro em vez do 'Analítico com Custo'
        for a in candidatos:
            if "custo" not in normalizar_nome_aba(a):
                return a
        return candidatos[0] if candidatos else None

    def identificar_aba_insumos(self) -> Optional[str]:
        """Identifica a aba de preços de insumos individuais (ISD/ICD/ISE).
        Prioriza ISD (sem desoneração) para correspondência com o padrão default.
        """
        for termo in ["isd", "icd", "ise"]:
            for a in self._xl.sheet_names:
                if normalizar_nome_aba(a) == termo:
                    return a
        return None

    def identificar_abas_precos(self) -> List[str]:
        """Identifica quais abas contêm dados de preços.
        """
        abas = self._xl.sheet_names
        if not abas:
            return []

        abas_precos = [
            a for a in abas
            if any(t in normalizar_nome_aba(a) for t in _TERMOS_PRECOS)
        ]

        if not abas_precos:
            abas_precos = [
                a for a in abas
                if not any(
                    ign in normalizar_nome_aba(a).upper()
                    for ign in _ABAS_IGNORADAS
                )
            ]

        return abas_precos

    # ------------------------------------------------------------------
    # Detecção de cabeçalho
    # ------------------------------------------------------------------

    def encontrar_header(self, df: pd.DataFrame) -> Optional[HeaderInfo]:
        """Detecta a linha de cabeçalho em um DataFrame de aba SINAPI.

        Estratégias (em ordem de prioridade):
          1. Linha única com estados + CODIGO/DESCRICAO na mesma linha.
          2. Duas linhas: linha com CODIGO/DESCRICAO (header) + linha
             anterior com siglas de estado (formato CSD/CCD/CSE).
          3. Fallback: linha com mais de 5 siglas de estado.

        Args:
            df: DataFrame lido com ``header=None``.

        Returns:
            HeaderInfo com as posições mapeadas, ou None se não encontrou.
        """
        # Estratégia 1: tudo na mesma linha
        header = self._buscar_header_completo(df)
        if header is not None:
            return header

        # Estratégia 2: header + estados na linha acima
        header = self._buscar_header_duas_linhas(df)
        if header is not None:
            return header

        # Estratégia 3: fallback apenas por estados
        return self._buscar_header_somente_estados(df)

    def _buscar_header_completo(self, df: pd.DataFrame) -> Optional[HeaderInfo]:
        """Busca header com estados + CODIGO/DESCRICAO nas primeiras 20 linhas."""
        for r_idx, row in df.head(20).iterrows():
            vals = [remover_acentos(str(v)).strip().upper() for v in row.values]

            tem_estados = any(v in COLUNAS_ESTADOS for v in vals)
            tem_cod = any("CODIGO" in v for v in vals)
            tem_desc = any("DESCRICAO" in v for v in vals)

            if tem_estados and (tem_cod or tem_desc):
                return self._montar_header_info(r_idx, vals)

        return None

    def _buscar_header_duas_linhas(self, df: pd.DataFrame) -> Optional[HeaderInfo]:
        """Busca header em duas linhas: CODIGO/DESCRICAO + estados acima.

        No formato CSD/CCD/CSE, a linha de cabeçalho (ex: row 9) tem
        Grupo, Código, Descrição, Unidade, Custo(R$), %AS... e a linha
        anterior (ex: row 8) contém as siglas dos estados nas colunas
        de Custo correspondentes.
        """
        for r_idx, row in df.head(20).iterrows():
            vals = [remover_acentos(str(v)).strip().upper() for v in row.values]

            tem_cod = any("CODIGO" in v for v in vals)
            tem_desc = any("DESCRICAO" in v for v in vals)

            if not (tem_cod or tem_desc):
                continue

            # Verificar se a linha anterior tem siglas de estado
            if r_idx < 1:
                continue

            prev_vals = [
                remover_acentos(str(v)).strip().upper()
                for v in df.iloc[r_idx - 1].values
            ]
            qtd_estados = sum(1 for v in prev_vals if v in COLUNAS_ESTADOS)

            if qtd_estados >= 5:
                # Montar header a partir da linha de labels
                info = self._montar_header_info(r_idx, vals)
                # Mapear estados a partir da linha acima
                for c_idx, val in enumerate(prev_vals):
                    if val in COLUNAS_ESTADOS:
                        info.mapa_estados[c_idx] = val.lower()
                return info

        return None

    def _buscar_header_somente_estados(self, df: pd.DataFrame) -> Optional[HeaderInfo]:
        """Busca header apenas por concentração de siglas de estado."""
        for r_idx, row in df.head(15).iterrows():
            vals = [remover_acentos(str(v)).strip().upper() for v in row.values]
            qtd_estados = sum(1 for v in vals if v in COLUNAS_ESTADOS)

            if qtd_estados > 5:
                return self._montar_header_info(r_idx, vals)

        return None

    @staticmethod
    def _montar_header_info(row_idx: int, vals: List[str]) -> HeaderInfo:
        """Monta HeaderInfo a partir dos valores normalizados de uma linha."""
        info = HeaderInfo(linha=row_idx)
        for c_idx, val in enumerate(vals):
            if val in COLUNAS_ESTADOS:
                info.mapa_estados[c_idx] = val.lower()
            elif "CODIGO" in val:
                info.col_codigo = c_idx
            elif "DESCRICAO" in val:
                info.col_descricao = c_idx
            elif "UNIDADE" in val:
                info.col_unidade = c_idx
            elif "GRUPO" in val:
                info.col_grupo = c_idx
        return info

    # ------------------------------------------------------------------
    # Classificação de tipo
    # ------------------------------------------------------------------

    def classificar_tipo_composicao(self, nome_aba: str) -> str:
        """Determina o tipo de composição pelo nome da aba.

        Args:
            nome_aba: Nome da aba da planilha.

        Returns:
            Um de: ``"Sem Desoneração"``, ``"Com Desoneração"``, ``"Empreitada"``.
        """
        aba_norm = normalizar_nome_aba(nome_aba).upper()
        for chave, tipo in _MAPA_TIPOS.items():
            if chave in aba_norm:
                return tipo
        return "Sem Desoneração"

    def extrair_metadados(self, sheet_hint: Optional[str] = None) -> Dict[str, Any]:
        """Extrai metadados específicos do formato SINAPI."""
        # Se não houver dica, tenta encontrar uma aba de dados
        aba_uso = sheet_hint
        if not aba_uso:
            abas = self.identificar_abas_dados()
            aba_uso = abas[0] if abas else self.sheet_names[0]

        df = self.ler_aba(aba_uso, header=None, nrows=20)
        
        mes_referencia = "UNKNOWN"
        desoneracao_raw = "UNKNOWN"
        uf = "BR"

        # Tentar extrair mês (Posição B3 no SINAPI)
        if len(df) > 2:
            val = str(df.iloc[2, 1]).strip()
            if len(val) > 4:
                mes_referencia = val

        # Tentar extrair desoneração (Posição D4 no SINAPI)
        if len(df) > 3 and len(df.columns) > 3:
            d = str(df.iloc[3, 3]).strip().upper()
            if d and d != "NAN":
                desoneracao_raw = d

        # Tentar extrair UF (D5 ou C3 no SINAPI)
        if len(df) > 4 and len(df.columns) > 3:
            u = str(df.iloc[4, 3]).strip().upper()
            if len(u) == 2 and u.isalpha():
                uf = u
        elif len(df) > 2 and len(df.columns) > 2:
            u = str(df.iloc[2, 2]).strip().upper()
            if len(u) == 2 and u.isalpha():
                uf = u

        return {
            "mes_referencia": mes_referencia,
            "uf": uf,
            "desoneracao": desoneracao_raw
        }

    # ------------------------------------------------------------------
    # Extração de registros
    # ------------------------------------------------------------------

    def extrair_registros_aba(
        self,
        nome_aba: str,
        mes_referencia: str,
    ) -> Tuple[List[dict], List[dict]]:
        """Extrai composições e preços de uma aba de preços.

        Args:
            nome_aba: Nome da aba a processar.
            mes_referencia: Mês de referência (ex: ``"09/2025"``).

        Returns:
            Tupla ``(composicoes, precos)`` onde cada elemento é uma lista
            de dicts prontos para upsert no banco.
        """
        df = self.ler_aba(nome_aba, header=None)
        header = self.encontrar_header(df)

        if header is None:
            logger.warning("Header não encontrado na aba '%s', pulando.", nome_aba)
            return [], []

        tipo_comp = self.classificar_tipo_composicao(nome_aba)
        df_dados = df.iloc[header.linha + 1:]

        # Detectar se os códigos são formulas HYPERLINK (pandas lê como 0)
        codigos_formula: Dict[int, str] = {}
        if header.col_codigo != -1:
            primeiros_codigos = df_dados.head(10).iloc[:, header.col_codigo]
            todos_zero = all(
                str(v).strip() in ('0', '0.0', '') or pd.isna(v)
                for v in primeiros_codigos
            )
            if todos_zero:
                logger.info(
                    "Detectados códigos zerados na aba '%s', extraindo de fórmulas.",
                    nome_aba,
                )
                codigos_formula = self._extrair_codigos_formula(
                    nome_aba, header.col_codigo, header.linha + 1,
                )

        composicoes: List[dict] = []
        precos: List[dict] = []

        for r_idx, row in df_dados.iterrows():
            # Usar código da fórmula se disponível
            # r_idx é o índice do DataFrame (0-based), openpyxl usa 1-based rows
            code_override = codigos_formula.get(r_idx)
            registro = self._processar_linha(
                row, header, mes_referencia, tipo_comp, code_override,
            )
            if registro is None:
                continue

            composicao, preco = registro
            composicoes.append(composicao)
            if preco is not None:
                precos.append(preco)

        return composicoes, precos

    def _extrair_codigos_formula(
        self,
        nome_aba: str,
        col_codigo: int,
        data_start_row: int,
    ) -> Dict[int, str]:
        """Extrai códigos reais de fórmulas HYPERLINK via openpyxl.

        Nas abas CSD/CCD/CSE, a coluna de código contém fórmulas como:
          =HYPERLINK("#"&..., 104658)
        O último argumento da fórmula é o código real.

        Args:
            nome_aba: Nome da aba.
            col_codigo: Índice da coluna de código (0-based).
            data_start_row: Linha de início dos dados no DataFrame (0-based).

        Returns:
            Dict mapeando índice do DataFrame (0-based) → código extraído.
        """
        wb = load_workbook(BytesIO(self._file_content), data_only=False, read_only=True)
        ws = wb[nome_aba]

        # Regex para extrair o último argumento numérico de HYPERLINK
        # Ex: =HYPERLINK("#"&CELL(...),104658) → captura '104658'
        padrao_hyperlink = re.compile(r'[,)]\s*(\d+)\s*\)\s*$')

        codigos: Dict[int, str] = {}
        # openpyxl col é 1-based
        col_opx = col_codigo + 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=0):
            if row_idx < data_start_row:
                continue
            cell = row[col_codigo] if col_codigo < len(row) else None
            if cell is None:
                continue

            val = cell.value
            if val is None:
                continue

            val_str = str(val)
            if val_str.startswith('='):
                match = padrao_hyperlink.search(val_str)
                if match:
                    codigos[row_idx] = match.group(1)
            elif val_str.strip() and val_str.strip() != '0':
                codigos[row_idx] = val_str.replace('.0', '').strip()

        wb.close()
        logger.info(
            "Extraídos %d códigos de fórmulas na aba '%s'.",
            len(codigos), nome_aba,
        )
        return codigos

    @staticmethod
    def _processar_linha(
        row: pd.Series,
        header: HeaderInfo,
        mes_referencia: str,
        tipo_composicao: str,
        code_override: Optional[str] = None,
    ) -> Optional[Tuple[dict, Optional[dict]]]:
        """Processa uma linha de dados e retorna composição + preço.

        Args:
            row: Linha do DataFrame.
            header: Informações do cabeçalho.
            mes_referencia: Mês de referência.
            tipo_composicao: Tipo (CSD/CCD/CSE).
            code_override: Código extraído de fórmula (se disponível).

        Returns:
            Tupla (composição, preço) ou None se a linha for inválida.
        """
        # Extrair código
        if code_override:
            cod = code_override
        else:
            cod_raw = row.iloc[header.col_codigo] if header.col_codigo != -1 else None
            if pd.isna(cod_raw) or str(cod_raw).strip() == "":
                return None
            cod = str(cod_raw).replace('.0', '').strip()

        if not cod.isdigit() and len(cod) < 3:
            return None

        # Extrair descrição e unidade
        desc_raw = row.iloc[header.col_descricao] if header.col_descricao != -1 else None
        desc = str(desc_raw).strip() if desc_raw is not None else ""
        unid = (
            str(row.iloc[header.col_unidade]).strip()
            if header.col_unidade != -1
            else "-"
        )

        # Extrair grupo
        grupo = "-"
        if header.col_grupo != -1:
            g_raw = row.iloc[header.col_grupo]
            if not pd.isna(g_raw) and str(g_raw).strip():
                grupo = str(g_raw).strip()

        # Composição
        composicao = {
            "codigo_composicao": cod,
            "descricao": desc,
            "unidade": unid,
            "grupo": grupo,
            "mes_referencia": mes_referencia,
        }

        # Preços por estado
        reg_preco = {
            "codigo_composicao": cod,
            "mes_referencia": mes_referencia,
            "tipo_composicao": tipo_composicao,
        }
        tem_valor = False
        for c_idx, sigla in header.mapa_estados.items():
            val = limpar_valor_moeda(row.iloc[c_idx])
            reg_preco[sigla] = val
            if val is not None:
                tem_valor = True
        
        preco = reg_preco if tem_valor else None
        return composicao, preco

    # ------------------------------------------------------------------
    # Extração da aba Analítico (hierarquia pai→filho)
    # ------------------------------------------------------------------

    def extrair_analitico(
        self,
        nome_aba: str,
        mes_referencia: str,
        fonte: str = "SINAPI",
    ) -> List[Dict[str, Any]]:
        """Extrai os relacionamentos composicao→insumo da aba Analítico.

        Estrutura da aba (row 9 = header):
          Grupo | Cód Composição | Tipo Item | Cód Item | Descrição | Unidade | Coeficiente | Situação

        Linhas onde 'Cód Item' é nulo são o cabeçalho da composição pai.
        Linhas com 'Tipo Item' = COMPOSICAO ou INSUMO são os filhos.

        Returns:
            Lista de dicts prontos para upsert em composicao_itens.
        """
        df = self.ler_aba(nome_aba, header=None, dtype=str)

        # --- Encontrar linha de cabeçalho ---
        col_cod_pai = -1
        col_tipo = -1
        col_cod_filho = -1
        col_desc = -1
        col_unid = -1
        col_coef = -1
        header_row = -1

        for r_idx, row in df.head(15).iterrows():
            vals = [remover_acentos(str(v)).strip().lower() for v in row.values]
            tem_codigo = any("codigo" in v and "composicao" in v for v in vals)
            tem_coef = any("coeficiente" in v for v in vals)
            if tem_codigo and tem_coef:
                header_row = r_idx
                for c_idx, v in enumerate(vals):
                    if "codigo" in v and "composicao" in v:
                        col_cod_pai = c_idx
                    elif "tipo" in v and "item" in v:
                        col_tipo = c_idx
                    elif "codigo" in v and "item" in v:
                        col_cod_filho = c_idx
                    elif "descricao" in v or "descri" in v:
                        col_desc = c_idx
                    elif "unidade" in v:
                        col_unid = c_idx
                    elif "coeficiente" in v:
                        col_coef = c_idx
                break

        if header_row == -1 or col_cod_pai == -1 or col_cod_filho == -1:
            logger.warning("Cabeçalho do Analítico não encontrado na aba '%s'.", nome_aba)
            return []

        registros: List[Dict[str, Any]] = []
        df_dados = df.iloc[header_row + 1:]

        for _, row in df_dados.iterrows():
            cod_pai_raw = row.iloc[col_cod_pai]
            cod_filho_raw = row.iloc[col_cod_filho]

            # Linhas de cabeçalho de composição: filho é nulo/nan
            if pd.isna(cod_filho_raw) or str(cod_filho_raw).strip() in ("", "None", "nan"):
                continue

            cod_pai = str(cod_pai_raw).replace('.0', '').strip()
            cod_filho = str(cod_filho_raw).replace('.0', '').strip()

            if not cod_pai.isdigit() or not cod_filho.isdigit():
                continue

            tipo_raw = row.iloc[col_tipo] if col_tipo != -1 else ""
            tipo = remover_acentos(str(tipo_raw)).strip().upper() if not pd.isna(tipo_raw) else ""

            desc_raw = row.iloc[col_desc] if col_desc != -1 else ""
            desc = str(desc_raw).strip() if not pd.isna(desc_raw) else ""

            unid_raw = row.iloc[col_unid] if col_unid != -1 else "-"
            unid = str(unid_raw).strip() if not pd.isna(unid_raw) else "-"

            coef_raw = row.iloc[col_coef] if col_coef != -1 else None
            try:
                coef = float(str(coef_raw).replace(',', '.')) if coef_raw and not pd.isna(coef_raw) else None
            except (ValueError, TypeError):
                coef = None

            if coef is None:
                continue

            registros.append({
                "codigo_pai": cod_pai,
                "codigo_filho": cod_filho,
                "quantidade_coeficiente": coef,
                "fonte": fonte,
                "mes_referencia": mes_referencia,
                "descricao_filho": desc,
                "unidade_filho": unid,
            })

        logger.info(
            "Aba '%s': %d relacionamentos analíticos extraídos.",
            nome_aba, len(registros)
        )
        return registros
