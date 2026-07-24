from typing import List, Dict, Any
from decimal import Decimal
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def gerar_planilha_orcamento_excel(orcamento: Dict[str, Any], itens: List[Dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orçamento Sintético"

    # Fontes e Estilos
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="001B3D", end_color="001B3D", fill_type="solid")
    font_bold = Font(name="Calibri", size=11, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    border_thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    # Cabeçalho do Projeto
    ws["A1"] = f"ORÇAMENTO: {orcamento.get('nome', '')}"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="001B3D")
    ws["A2"] = f"Cliente: {orcamento.get('cliente', '')} | Base: {orcamento.get('base_referencia', '')} | BDI: {orcamento.get('bdi', 0)}%"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="64748B")

    # Linhas da Tabela
    headers = ["Item", "Código", "Descrição da Composição", "Unid.", "Qtd.", "Preço Unit. (R$)", "Preço Total (R$)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    row_num = 5
    total_geral = Decimal("0.0")

    for idx, item in enumerate(itens, 1):
        qtd = Decimal(str(item.get("quantidade", 0)))
        pu = Decimal(str(item.get("preco_unitario") or 0))
        pt = Decimal(str(item.get("preco_total") or (qtd * pu)))
        total_geral += pt

        ws.cell(row=row_num, column=1, value=idx).alignment = align_center
        ws.cell(row=row_num, column=2, value=item.get("codigo_composicao", "")).alignment = align_center
        ws.cell(row=row_num, column=3, value=item.get("descricao", ""))
        ws.cell(row=row_num, column=4, value=item.get("unidade", "")).alignment = align_center
        
        cell_qtd = ws.cell(row=row_num, column=5, value=float(qtd))
        cell_qtd.number_format = "#,##0.00"
        cell_qtd.alignment = align_right

        cell_pu = ws.cell(row=row_num, column=6, value=float(pu))
        cell_pu.number_format = "R$ #,##0.00"
        cell_pu.alignment = align_right

        cell_pt = ws.cell(row=row_num, column=7, value=float(pt))
        cell_pt.number_format = "R$ #,##0.00"
        cell_pt.alignment = align_right

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).border = border_thin

        row_num += 1

    # Linha de Total Geral
    ws.cell(row=row_num, column=3, value="TOTAL GERAL DO ORÇAMENTO").font = font_bold
    ws.cell(row=row_num, column=3).alignment = align_right
    cell_tot = ws.cell(row=row_num, column=7, value=float(total_geral))
    cell_tot.font = font_bold
    cell_tot.number_format = "R$ #,##0.00"
    cell_tot.alignment = align_right

    # Largura das colunas
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
